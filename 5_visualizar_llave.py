"""
==============================================================
PREDICCIÓN MUNDIALISTA 2026 - Script 5: Visualizar Llave
==============================================================
Genera dos outputs principales:

1. Bracket visual en matplotlib:
   - Fondo degradado azul noche #0A1628
   - Seis columnas: Dieciseisavos de Final → Octavos de Final → Cuartos →
     Semifinales → Final → Campeón
   - Tercer y Cuarto Puesto debajo de semifinales
   - Cajas FancyBboxPatch con bordes redondeados
   - Líneas Bezier conectando ganadores
   - Panel lateral con Top 5 favoritos
   - Exportar en 300 DPI como salida/bracket_mundial_2026.png

2. Excel salida/predicciones_llave_2026.xlsx con 4 hojas:
   - "Predicciones": tabla completa
   - "Resumen por Ronda": partidos y goles por ronda
   - "Dashboard": Top 5 + gráficos incrustados
   - "Campeón": trofeo ASCII + nombre del campeón
==============================================================
"""

import os
import sys
import pickle
import warnings
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch, PathPatch
    from matplotlib.path import Path
    HAS_MPL = True
except ImportError:
    HAS_MPL = False
    print("  ⚠ matplotlib no disponible. No se generará el bracket PNG.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    print("  ⚠ openpyxl no disponible. No se generará el Excel.")
    print("    → Instale con: pip install openpyxl")

# ─── Rutas ────────────────────────────────────────────────────
RAIZ = os.path.dirname(os.path.abspath(__file__))
PROCESADOS = os.path.join(RAIZ, "procesados")
SALIDA = os.path.join(RAIZ, "salida")
os.makedirs(SALIDA, exist_ok=True)

FLAG_MAP = {}


def flag(equipo):
    return FLAG_MAP.get(str(equipo).strip().lower(), "")


def cargar_pickle(nombre):
    ruta = os.path.join(PROCESADOS, f"{nombre}.pkl")
    if not os.path.exists(ruta):
        print(f"  ⚠ {nombre}.pkl no encontrado.")
        return None
    with open(ruta, "rb") as f:
        return pickle.load(f)


# ══════════════════════════════════════════════════════════════
#  INICIO
# ══════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  PREDICCIÓN MUNDIALISTA 2026 - VISUALIZACIÓN DE LLAVE")
print("═"*60)

# ─── Cargar datos ─────────────────────────────────────────────
print("\n[1/3] Cargando predicciones...")
datos = cargar_pickle("predicciones_llave")

if datos is None:
    print("  ✗ Error: predicciones_llave.pkl no encontrado.")
    print("    → Ejecute primero: python 4_simular_llave.py")
    sys.exit(1)

df = datos["resultados"]
campeon = datos.get("campeon", "Por determinar")
ruta_prob = datos.get("ruta_probabilidades", {})
favoritos = datos.get("favoritos", [])

print(f"  ✓ {len(df)} partidos cargados")
print(f"  ✓ Campeón: {campeon}")


# ══════════════════════════════════════════════════════════════
#  PARTE 1: Bracket visual con matplotlib
# ══════════════════════════════════════════════════════════════
if HAS_MPL:
    print("\n[2/3] Generando bracket visual...")

    # Nomenclatura oficial en español, sin saltarse rondas:
    #   Dieciseisavos (32) → Octavos (16) → Cuartos (8) → Semifinales (4) → Final (2)
    rondas_nombres = [
        "Dieciseisavos de Final",
        "Octavos de Final",
        "Cuartos de Final",
        "Semifinales",
        "Final",
    ]

    # Configurar figura
    fig_width = 22
    fig_height = 14
    fig, ax = plt.subplots(1, 1, figsize=(fig_width, fig_height))
    ax.set_facecolor("#0A1628")
    fig.patch.set_facecolor("#0A1628")
    ax.set_xlim(0, 100)
    ax.set_ylim(0, 100)
    ax.axis("off")

    # Colores
    COLOR_BORDE = "#10b981"
    COLOR_BORDE_GANADOR = "#eab308"
    COLOR_TEXTO = "#f0f0f0"
    COLOR_GANADOR = "#fcd34d"
    COLOR_PERDEDOR = "#6b7280"
    COLOR_FONDO_CAJA = "#111827"
    COLOR_LINEA = "#4ade80"

    # Parámetros de layout
    n_rondas = len(rondas_nombres)
    col_width = 15
    col_start = 3
    col_spacing = 17

    # Título
    ax.text(50, 97, "PREDICCIÓN MUNDIALISTA 2026", fontsize=22, fontweight="bold",
            color="#fcd34d", ha="center", va="top",
            fontfamily="sans-serif")

    # Dibujar encabezados de columnas
    for i, nombre in enumerate(rondas_nombres):
        x = col_start + i * col_spacing
        ax.text(x + col_width/2, 93, nombre.upper(), fontsize=9, fontweight="bold",
                color="#64748b", ha="center", va="top",
                fontfamily="sans-serif")

    # Campeón
    ax.text(col_start + 5 * col_spacing + col_width/2, 93, "CAMPEÓN",
            fontsize=11, fontweight="bold", color="#eab308", ha="center", va="top")

    # Dibujar cajas de partidos por ronda
    cajas_posiciones = {}  # ronda_idx -> [(x, y_centro, equipo_local, equipo_visitante, gl, gv, ganador)]

    for r_idx, ronda_nombre in enumerate(rondas_nombres):
        partidos_ronda = df[df["Ronda"] == ronda_nombre]
        x = col_start + r_idx * col_spacing
        n_partidos = len(partidos_ronda)

        if n_partidos == 0:
            continue

        box_height = 7
        gap = max(2, (85 - n_partidos * (box_height + 2)) / max(n_partidos - 1, 1))
        total_height = n_partidos * box_height + (n_partidos - 1) * gap
        y_start = 88 - (88 - total_height) / 2

        cajas_posiciones[r_idx] = []

        for p_idx, (_, partido) in enumerate(partidos_ronda.iterrows()):
            y_top = y_start - p_idx * (box_height + gap)
            y_center = y_top - box_height / 2

            local = partido["Equipo Local"]
            visitante = partido["Equipo Visitante"]
            gl = partido["Goles Local"]
            gv = partido["Goles Visitante"]
            ganador = partido["Ganador"]

            # Caja con bordes redondeados
            border_color = COLOR_BORDE_GANADOR if r_idx >= 3 else COLOR_BORDE
            bbox = FancyBboxPatch(
                (x, y_top - box_height), col_width, box_height,
                boxstyle="round,pad=0.3",
                facecolor=COLOR_FONDO_CAJA,
                edgecolor=border_color,
                linewidth=2.0,
                zorder=3
            )
            ax.add_patch(bbox)

            # Texto de equipos
            color_l = COLOR_GANADOR if ganador == local else COLOR_PERDEDOR
            color_v = COLOR_GANADOR if ganador == visitante else COLOR_PERDEDOR
            fontw_l = "bold" if ganador == local else "normal"
            fontw_v = "bold" if ganador == visitante else "normal"

            fl = flag(local)
            fv = flag(visitante)

            # Barra de probabilidad
            prob_l = partido["Prob Local"]
            bar_width_l = max(1, col_width * 0.3 * prob_l * 2)
            bar_rect = plt.Rectangle(
                (x + 0.5, y_top - 3), bar_width_l, 1,
                facecolor="#10b981", alpha=0.3, zorder=2
            )
            ax.add_patch(bar_rect)

            ax.text(x + 0.8, y_top - 1.5, f"{fl} {local[:12]}", fontsize=7,
                    color=color_l, fontweight=fontw_l, va="center", zorder=4)
            ax.text(x + 0.8, y_top - 4.5, f"{fv} {visitante[:12]}", fontsize=7,
                    color=color_v, fontweight=fontw_v, va="center", zorder=4)

            # Marcador
            ax.text(x + col_width - 1.5, y_top - 3, f"{gl}-{gv}", fontsize=8,
                    color=COLOR_TEXTO, fontweight="bold", ha="right", va="center", zorder=4)

            cajas_posiciones[r_idx].append({
                "x": x, "y_center": y_center,
                "x_right": x + col_width,
                "local": local, "visitante": visitante,
                "ganador": ganador, "gl": gl, "gv": gv
            })

    # Dibujar líneas Bezier conectando rondas
    for r_idx in range(len(rondas_nombres) - 1):
        if r_idx not in cajas_posiciones or (r_idx + 1) not in cajas_posiciones:
            continue
        cajas_origen = cajas_posiciones[r_idx]
        cajas_destino = cajas_posiciones[r_idx + 1]

        # Emparejar: cada 2 cajas de origen van a 1 caja de destino
        for d_idx, dest in enumerate(cajas_destino):
            origen_idx_1 = min(d_idx * 2, len(cajas_origen) - 1)
            origen_idx_2 = min(d_idx * 2 + 1, len(cajas_origen) - 1)

            for o_idx in [origen_idx_1, origen_idx_2]:
                if o_idx >= len(cajas_origen):
                    continue
                orig = cajas_posiciones[r_idx][o_idx]
                x1 = orig["x_right"]
                y1 = orig["y_center"]
                x2 = dest["x"]
                y2 = dest["y_center"]

                # Curva Bezier
                x_mid = (x1 + x2) / 2
                bezier_path = Path(
                    [(x1, y1), (x_mid, y1), (x_mid, y2), (x2, y2)],
                    [Path.MOVETO, Path.CURVE4, Path.CURVE4, Path.CURVE4]
                )
                patch = PathPatch(
                    bezier_path,
                    facecolor="none",
                    edgecolor=COLOR_LINEA,
                    lw=1.5,
                    alpha=0.5,
                    zorder=1
                )
                ax.add_patch(patch)

    # Campeón — círculo dorado
    champ_x = col_start + 5 * col_spacing + col_width / 2
    champ_y = 55
    circle = plt.Circle((champ_x, champ_y), 6, facecolor="#052e16",
                         edgecolor="#eab308", linewidth=4, zorder=5)
    ax.add_patch(circle)
    ax.text(champ_x, champ_y + 2, "COPA", fontsize=14, fontweight="bold",
            color="#fcd34d", ha="center", va="center", zorder=6)
    ax.text(champ_x, champ_y - 3, campeon[:12], fontsize=9, fontweight="bold",
            color="#fcd34d", ha="center", va="center", zorder=6)

    # Tercer y Cuarto Puesto
    tercer = df[df["Ronda"] == "Tercer y Cuarto Puesto"]
    if not tercer.empty:
        t = tercer.iloc[0]
        tercero_eq = t["Ganador"]
        ax.text(col_start + 3 * col_spacing + col_width/2, 5,
                f"3° {flag(tercero_eq)} {tercero_eq}",
                fontsize=8, color="#64748b", ha="center", va="bottom")

    # Panel lateral: Top 5 favoritos
    panel_x = 88
    panel_y_start = 82
    ax.text(panel_x, panel_y_start, "TOP 5", fontsize=12, fontweight="bold",
            color="#eab308", ha="center", va="top")

    colores_top = ["#fcd34d", "#a78bfa", "#38bdf8", "#4ade80", "#fb923c"]

    for i, (eq, probs) in enumerate(favoritos[:5]):
        y = panel_y_start - 4 - i * 14
        f = flag(eq)
        ax.text(panel_x, y, f"{f} {eq[:14]}", fontsize=8, fontweight="bold",
                color=colores_top[i], ha="center", va="top")

        # Barras de probabilidad
        for j, (instancia, color) in enumerate([
            ("Campeón", "#fcd34d"), ("Final", "#a78bfa"),
            ("Semifinales", "#38bdf8"), ("Cuartos", "#4ade80")
        ]):
            prob_val = probs.get(instancia, 0)
            bar_w = max(0, prob_val * 10)
            bar_y = y - 4 - j * 2.5
            ax.add_patch(plt.Rectangle(
                (panel_x - 5, bar_y), bar_w, 1.5,
                facecolor=color, alpha=0.6, zorder=2
            ))
            if prob_val > 0.01:
                ax.text(panel_x - 5 + bar_w + 0.3, bar_y + 0.75,
                        f"{prob_val:.0%}", fontsize=5, color=color, va="center")

    # Guardar
    ruta_png = os.path.join(SALIDA, "bracket_mundial_2026.png")
    plt.savefig(ruta_png, dpi=300, bbox_inches="tight",
                facecolor="#0A1628", edgecolor="none")
    plt.close()
    print(f"  ✓ bracket_mundial_2026.png guardado (300 DPI)")

else:
    print("\n[2/3] Omitiendo bracket visual (matplotlib no disponible)")


# ══════════════════════════════════════════════════════════════
#  PARTE 2: Excel con 4 hojas
# ══════════════════════════════════════════════════════════════
if HAS_XLSX:
    print("\n[3/3] Generando Excel con 4 hojas...")

    ruta_xlsx = os.path.join(SALIDA, "predicciones_llave_2026.xlsx")

    # Estilos
    fill_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    fill_alt = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fill_gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    fill_dark_green = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_gold = Font(name="Calibri", bold=True, color="1B5E20", size=11)
    font_campeon = Font(name="Calibri", bold=True, color="1B5E20", size=24)
    font_trofeo = Font(name="Consolas", bold=True, color="FFD700", size=14)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="1B5E20"),
        right=Side(style="thin", color="1B5E20"),
        top=Side(style="thin", color="1B5E20"),
        bottom=Side(style="thin", color="1B5E20")
    )

    with pd.ExcelWriter(ruta_xlsx, engine="openpyxl") as writer:
        # ─── HOJA 1: Predicciones ──────────────────────────────
        cols_pred = ["Ronda", "Match", "Equipo Local", "Equipo Visitante",
                     "Goles Local", "Goles Visitante", "Ganador",
                     "Prob Local", "Prob Visitante", "Prob Victoria",
                     "RF Local", "RF Empate", "RF Visitante",
                     "XGB Local", "XGB Empate", "XGB Visitante",
                     "SVM Local", "SVM Empate", "SVM Visitante",
                     "Stack Local", "Stack Empate", "Stack Visitante",
                     "Poisson λ Local", "Poisson λ Visitante"]

        cols_disponibles = [c for c in cols_pred if c in df.columns]
        df_export = df[cols_disponibles].copy()
        df_export.to_excel(writer, sheet_name="Predicciones", index=False)

        ws = writer.sheets["Predicciones"]

        # Formatear encabezados
        for col_idx in range(1, len(cols_disponibles) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = thin_border

        # Filas alternas
        for row_idx in range(2, len(df_export) + 2):
            for col_idx in range(1, len(cols_disponibles) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_idx % 2 == 0:
                    cell.fill = fill_alt
                cell.border = thin_border
                cell.alignment = align_center

        # Fila de la Final del campeón en dorado
        for row_idx in range(2, len(df_export) + 2):
            ronda_val = ws.cell(row=row_idx, column=1).value
            if ronda_val and "Final" in str(ronda_val) and "Cuartos" not in str(ronda_val):
                for col_idx in range(1, len(cols_disponibles) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_gold
                    cell.font = font_gold

        # Ajustar anchos
        for col_idx in range(1, len(cols_disponibles) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        # ─── HOJA 2: Resumen por Ronda ────────────────────────
        resumen = df.groupby("Ronda").agg(
            Partidos=("Ronda", "count"),
            Goles_Totales=("Goles Local", lambda x: (x + df.loc[x.index, "Goles Visitante"]).sum()),
        ).reset_index()
        resumen["Promedio Goles"] = (resumen["Goles_Totales"] / resumen["Partidos"]).round(2)

        resumen.to_excel(writer, sheet_name="Resumen por Ronda", index=False)
        ws2 = writer.sheets["Resumen por Ronda"]

        for col_idx in range(1, len(resumen.columns) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center

        for col_idx in range(1, len(resumen.columns) + 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 22

        # ─── HOJA 3: Dashboard ────────────────────────────────
        # Top 5 favoritos
        dash_data = []
        for eq, probs in favoritos[:5]:
            dash_data.append({
                "Equipo": eq,
                "Prob Campeón": round(probs.get("Campeón", 0), 4),
                "Prob Final": round(probs.get("Final", 0), 4),
                "Prob Semifinal": round(probs.get("Semifinales", 0), 4),
                "Prob Cuartos": round(probs.get("Cuartos", 0), 4),
                "Prob 16avos": round(probs.get("16avos", 0), 4),
            })

        df_dash = pd.DataFrame(dash_data)
        df_dash.to_excel(writer, sheet_name="Dashboard", index=False, startrow=2)
        ws3 = writer.sheets["Dashboard"]

        # Título con degradado simulado
        ws3.merge_cells("A1:F1")
        title_cell = ws3.cell(row=1, column=1)
        title_cell.value = "DASHBOARD — TOP 5 FAVORITOS MUNDIAL 2026"
        title_cell.fill = fill_dark_green
        title_cell.font = Font(name="Calibri", bold=True, color="4CAF50", size=16)
        title_cell.alignment = align_center

        ws3.merge_cells("A2:F2")
        subtitle = ws3.cell(row=2, column=1)
        subtitle.value = "Probabilidades acumuladas de llegar a cada instancia"
        subtitle.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        subtitle.font = Font(name="Calibri", italic=True, color="81C784", size=10)
        subtitle.alignment = align_center

        # Encabezados de tabla
        for col_idx in range(1, len(df_dash.columns) + 1):
            cell = ws3.cell(row=3, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center

        # Formato condicional de escala de colores verde
        from openpyxl.formatting.rule import ColorScaleRule
        green_scale = ColorScaleRule(
            start_type="min", start_color="E8F5E9",
            mid_type="percentile", mid_value=50, mid_color="66BB6A",
            end_type="max", end_color="1B5E20"
        )
        ws3.conditional_formatting.add(
            f"B4:F{3 + len(df_dash)}", green_scale
        )

        for col_idx in range(1, len(df_dash.columns) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 18

        # Gráfico 1: Barras horizontales de probabilidad de ser campeón
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.title = "Probabilidad de ser Campeón"
        chart1.x_axis.title = "Probabilidad"
        chart1.y_axis.title = ""
        chart1.style = 10

        data1 = Reference(ws3, min_col=2, min_row=3, max_row=3 + len(df_dash))
        cats1 = Reference(ws3, min_col=1, min_row=4, max_row=3 + len(df_dash))
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.shape = 4
        chart1.width = 18
        chart1.height = 10

        ws3.add_chart(chart1, f"A{5 + len(df_dash)}")

        # Gráfico 2: Gráfico apilado por ronda para los 5 favoritos
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.grouping = "stacked"
        chart2.title = "Probabilidades por Instancia (Top 5)"
        chart2.style = 10
        chart2.width = 18
        chart2.height = 10

        data2 = Reference(ws3, min_col=2, min_row=3, max_col=6, max_row=3 + len(df_dash))
        cats2 = Reference(ws3, min_col=1, min_row=4, max_row=3 + len(df_dash))
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)

        ws3.add_chart(chart2, f"A{17 + len(df_dash)}")

        # ─── HOJA 4: Campeón ──────────────────────────────────
        ws4 = writer.sheets.create("Campeón")
        writer.sheets["Campeón"] = ws4

        # Trofeo ASCII en celda A1
        trofeo_ascii = """  ╔══════════════════════════╗
  ║                          ║
  ║        CAMPEÓN          ║
  ║                          ║
  ╚══════════════════════════╝"""

        ws4.merge_cells("A1:F8")
        cell_trofeo = ws4.cell(row=1, column=1)
        cell_trofeo.value = trofeo_ascii
        cell_trofeo.font = font_trofeo
        cell_trofeo.fill = fill_dark_green
        cell_trofeo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Nombre del campeón
        ws4.merge_cells("A10:F12")
        cell_campeon = ws4.cell(row=10, column=1)
        fc = flag(campeon)
        cell_campeon.value = f"{fc} {campeon}"
        cell_campeon.font = font_campeon
        cell_campeon.fill = fill_dark_green
        cell_campeon.alignment = Alignment(horizontal="center", vertical="center")

        # Info adicional
        ws4.merge_cells("A14:F14")
        cell_info = ws4.cell(row=14, column=1)
        cell_info.value = f"Probabilidad en la Final: {datos.get('prob_campeon', 0):.0%}"
        cell_info.font = Font(name="Calibri", size=12, color="81C784")
        cell_info.fill = fill_dark_green
        cell_info.alignment = Alignment(horizontal="center")

        for col in range(1, 7):
            ws4.column_dimensions[get_column_letter(col)].width = 14

    print(f"  ✓ predicciones_llave_2026.xlsx guardado con 4 hojas")
    print(f"    • Predicciones: {len(df_export)} partidos con probabilidades por modelo")
    print(f"    • Resumen por Ronda: estadísticas por ronda")
    print(f"    • Dashboard: Top 5 favoritos + 2 gráficos")
    print(f"    • Campeón: trofeo ASCII + {campeon}")

else:
    print("\n[3/3] Omitiendo Excel (openpyxl no disponible)")


# ─── Final ─────────────────────────────────────────────────────
print(f"\n{'═'*60}")
print("  VISUALIZACIÓN COMPLETADA")
print(f"  Archivos generados en salida/:")
if HAS_MPL:
    print(f"    • bracket_mundial_2026.png (300 DPI)")
if HAS_XLSX:
    print(f"    • predicciones_llave_2026.xlsx (4 hojas)")
print("═"*60 + "\n")
